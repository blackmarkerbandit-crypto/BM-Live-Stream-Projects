import openpyxl
import re

filepath = r'd:\Dropbox\WORK FILES\BMB\Business Corp Projects\STREAMING SETUP 2 - 2026\Setup 2 Pricing and Purchase Sheet 2026.xlsx'

print("Loading workbook...", flush=True)
wb = openpyxl.load_workbook(filepath)
ws = wb['GAME PLAN (3)']

print("Checking and fixing column A formulas (Loss/Gain = C-B per row)...", flush=True)

# After the row deletion, rows 12-33 in column A have stale formulas.
# The correct formula for row N in column A should be =CN-BN
# Let's check each row and fix where formula doesn't match row number

issues_fixed = 0
for row in ws.iter_rows(min_row=4, max_row=ws.max_row - 1, min_col=1, max_col=1):
    for cell in row:
        r = cell.row
        expected = f'=C{r}-B{r}'
        current = cell.value
        if current != expected:
            print(f"  Row {r}: formula is {repr(current)}, should be {repr(expected)} -- FIXING", flush=True)
            cell.value = expected
            issues_fixed += 1
        else:
            print(f"  Row {r}: formula OK: {repr(current)}", flush=True)

print(f"\nFixed {issues_fixed} formula(s).", flush=True)

# Now check the subtotal row (row 19) - SUM(B4:B19) would include itself,
# which is wrong. Let's check what row 19 column B has.
print(f"\nRow 19 B: {repr(ws['B19'].value)}", flush=True)
print(f"Row 19 D: {repr(ws['D19'].value)}", flush=True)

# The subtotal row was row 20 originally (SUM B4:B19), and after deleting row 12,
# it shifted to row 19. It still says =SUM(B4:B19) which would now include itself.
# It should be =SUM(B4:B18) since the data above ends at row 18.
# Let's verify: last data row before the subtotal is row 18.
print(f"\nRow 18 D: {repr(ws['D18'].value)}, B18={ws['B18'].value}", flush=True)
print(f"Row 17 D: {repr(ws['D17'].value)}, B17={ws['B17'].value}", flush=True)

# Fix subtotal formula if it references itself
subtotal_row = 19
subtotal_formula = ws.cell(subtotal_row, 2).value
print(f"\nSubtotal row {subtotal_row} B: {repr(subtotal_formula)}", flush=True)

if subtotal_formula == '=SUM(B4:B19)':
    print("  Subtotal formula includes itself! Fixing to =SUM(B4:B18)...", flush=True)
    ws.cell(subtotal_row, 2).value = '=SUM(B4:B18)'
    print(f"  Fixed to =SUM(B4:B18)", flush=True)

# Check TOTAL row
total_row = ws.max_row
print(f"\nTOTAL row {total_row}:", flush=True)
print(f"  A{total_row}: {repr(ws.cell(total_row, 1).value)}", flush=True)
print(f"  B{total_row}: {repr(ws.cell(total_row, 2).value)}", flush=True)
print(f"  C{total_row}: {repr(ws.cell(total_row, 3).value)}", flush=True)

# The TOTAL row formulas reference specific row ranges - let's check them.
# Original (before delete): A34=SUM(A21:A33)+A20, B34=SUM(B21:B33)+B20
# After delete: TOTAL is row 33. Formulas should now be:
# A33=SUM(A20:A32)+A19, B33=SUM(B20:B32)+B19
# But openpyxl may have already shifted them or not.
# Let's check and fix if needed.

# The subtotal is at row 19 (was row 20 before delete).
# The section below subtotal starts at row 20 (was row 21 before delete).
# TOTAL is at row 33 (was row 34 before delete).

expected_a_total = f'=SUM(A20:A32)+A19'
expected_b_total = f'=SUM(B20:B32)+B19'
expected_c_total = f'=SUM(C4:C32)'

current_a = ws.cell(total_row, 1).value
current_b = ws.cell(total_row, 2).value
current_c = ws.cell(total_row, 3).value

print(f"\nChecking TOTAL formulas...", flush=True)

# Fix A total
if current_a != expected_a_total:
    print(f"  A{total_row}: {repr(current_a)} -> {repr(expected_a_total)}", flush=True)
    ws.cell(total_row, 1).value = expected_a_total

# Fix B total
if current_b != expected_b_total:
    print(f"  B{total_row}: {repr(current_b)} -> {repr(expected_b_total)}", flush=True)
    ws.cell(total_row, 2).value = expected_b_total

# Fix C total
if current_c != expected_c_total:
    print(f"  C{total_row}: {repr(current_c)} -> {repr(expected_c_total)}", flush=True)
    ws.cell(total_row, 3).value = expected_c_total

# Print final full state
print("\n--- FINAL STATE ---", flush=True)
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f'Row {cell.row}, Col {cell.column_letter}: {repr(cell.value)}', flush=True)

print("\nSaving...", flush=True)
wb.save(filepath)
print("Saved successfully!", flush=True)
