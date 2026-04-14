import openpyxl
import sys

print("openpyxl imported OK", flush=True)

filepath = r'd:\Dropbox\WORK FILES\BMB\Business Corp Projects\STREAMING SETUP 2 - 2026\Setup 2 Pricing and Purchase Sheet 2026.xlsx'

wb = openpyxl.load_workbook(filepath)
print('Sheets:', wb.sheetnames, flush=True)

ws = wb['GAME PLAN (3)']
print('Dimensions:', ws.dimensions, flush=True)
print('Max row:', ws.max_row, 'Max col:', ws.max_column, flush=True)
print(flush=True)

for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f'Row {cell.row}, Col {cell.column} ({cell.column_letter}): {repr(cell.value)}', flush=True)

print("DONE", flush=True)
