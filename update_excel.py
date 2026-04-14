import openpyxl
from openpyxl import load_workbook
from copy import copy

filepath = r'd:\Dropbox\WORK FILES\BMB\Business Corp Projects\STREAMING SETUP 2 - 2026\Setup 2 Pricing and Purchase Sheet 2026.xlsx'

print("Loading workbook...", flush=True)
wb = load_workbook(filepath)
ws = wb['GAME PLAN (3)']

print("Sheet loaded. Starting changes...", flush=True)

# -----------------------------------------------------------------------
# STEP 1: Update Row 4 — SKB rack case → Gator Cases
# -----------------------------------------------------------------------
print("\nStep 1: Updating Row 4 (SKB to Gator Cases)...", flush=True)
print(f"  Before: D4={repr(ws['D4'].value)}, B4={ws['B4'].value}, E4={repr(ws['E4'].value)}", flush=True)

ws['D4'].value = 'Gator Cases GR-12L-LOCK 12U Mold Shallow Rolling Rack'
ws['B4'].value = 379.99
ws['E4'].value = 'https://www.amazon.com/Gator-Cases-GR-12L-LOCK-Rolling/dp/B0002M64BM'
# A4 already has formula =C4-B4, which is still correct

print(f"  After:  D4={repr(ws['D4'].value)}, B4={ws['B4'].value}, E4={repr(ws['E4'].value)}", flush=True)

# -----------------------------------------------------------------------
# STEP 2: Delete Row 12 — MR NET Router (Upgrade) at $699
# -----------------------------------------------------------------------
print("\nStep 2: Deleting Row 12 (MR NET Router Upgrade)...", flush=True)
print(f"  Row 12 before delete: D12={repr(ws['D12'].value)}, B12={ws['B12'].value}", flush=True)

ws.delete_rows(12)

# After deletion, what was row 13 is now row 12, row 16 is now row 15, row 20 is now row 19, row 34 is now row 33
print(f"  Row 12 after delete (was row 13): D12={repr(ws['D12'].value)}", flush=True)
print(f"  New max row: {ws.max_row}", flush=True)

# -----------------------------------------------------------------------
# STEP 3: Update the ATEM row — now row 15 (was row 16)
# -----------------------------------------------------------------------
print("\nStep 3: Updating ATEM row (now row 15)...", flush=True)
# Find the ATEM row by scanning (to be safe after deletion)
atem_row = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value and 'ATEM SDI Extreme ISO' in str(cell.value):
            atem_row = cell.row
            break
    if atem_row:
        break

if atem_row is None:
    print("  ERROR: Could not find ATEM row!", flush=True)
else:
    print(f"  Found ATEM row at row {atem_row}: D={repr(ws.cell(atem_row, 4).value)}, B={ws.cell(atem_row, 2).value}", flush=True)
    ws.cell(atem_row, 4).value = 'Blackmagic Design ATEM SDI Pro ISO Switcher'
    ws.cell(atem_row, 2).value = 995.00
    ws.cell(atem_row, 5).value = 'https://www.bhphotovideo.com/c/product/1627626-REG/blackmagic_design_swatemsdipro4k_atem_sdi_pro_iso.html'
    print(f"  After:  D{atem_row}={repr(ws.cell(atem_row, 4).value)}, B{atem_row}={ws.cell(atem_row, 2).value}", flush=True)

# -----------------------------------------------------------------------
# STEP 4: Verify / fix TOTAL row formulas
# -----------------------------------------------------------------------
print("\nStep 4: Checking TOTAL row...", flush=True)
# Find the TOTAL row
total_row = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value == 'TOTAL':
            total_row = cell.row
            break
    if total_row:
        break

if total_row is None:
    print("  ERROR: Could not find TOTAL row!", flush=True)
else:
    print(f"  TOTAL row is at row {total_row}", flush=True)
    print(f"  A{total_row}={repr(ws.cell(total_row, 1).value)}", flush=True)
    print(f"  B{total_row}={repr(ws.cell(total_row, 2).value)}", flush=True)
    print(f"  C{total_row}={repr(ws.cell(total_row, 3).value)}", flush=True)

# Also find the subtotal row (SUM(B4:B19) type row)
print("\nLooking for subtotal row (SUM of B4:Bxx)...", flush=True)
for row in ws.iter_rows(min_col=2, max_col=2):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and 'SUM' in cell.value.upper():
            print(f"  Row {cell.row}, Col B: {repr(cell.value)}", flush=True)

# Print all non-empty cells to verify final state
print("\n--- FINAL STATE OF SHEET ---", flush=True)
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            print(f'Row {cell.row}, Col {cell.column} ({cell.column_letter}): {repr(cell.value)}', flush=True)

# -----------------------------------------------------------------------
# SAVE
# -----------------------------------------------------------------------
print("\nSaving workbook...", flush=True)
wb.save(filepath)
print("Saved successfully!", flush=True)
